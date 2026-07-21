"""
TurFresh - Alerta Semanal de GSC
==================================
Motor de comparacao contra media movel de 4 semanas, com piso de volume
em CADA semana (nao so na media) para evitar ruido de numeros pequenos.

Por que piso por semana e nao so na media: se 3 das 4 semanas anteriores
tiveram 20 impressoes e uma teve 1.200 (um pico isolado, ex. trafego de
bot ou evento sazonal), a MEDIA pode passar de 300 mesmo que a pagina seja
pequena no dia a dia. Exigir volume minimo em pelo menos 3 das 4 semanas
individualmente evita que um outlier valide um piso que nao existe de verdade.

Este primeiro corte tem so o Gatilho 1 (vazamento de CTR) funcionando.
Os outros 5 entram depois, um de cada vez, em cima desse mesmo motor.
"""

import os
import re
from collections import defaultdict
from datetime import date, timedelta

import pandas as pd
from google.oauth2 import service_account
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from google.analytics.data_v1beta import BetaAnalyticsDataClient
    from google.analytics.data_v1beta.types import (
        DateRange, Dimension, Metric, Filter, FilterExpression, RunReportRequest,
    )
    GA4_AVAILABLE = True
except ImportError:
    GA4_AVAILABLE = False

GA4_PROPERTY_ID = os.environ.get("GA4_PROPERTY_ID", "")

# ===========================================================================
# CONFIG
# ===========================================================================
SITE_URL = os.environ.get("SITE_URL", "https://turfresh.com/")
GSC_CLIENT_EMAIL = os.environ.get("GSC_CLIENT_EMAIL")
GSC_PRIVATE_KEY = os.environ.get("GSC_PRIVATE_KEY")

WEEK_LAG_DAYS = 3          # GSC leva ~3 dias para fechar os dados de uma semana
N_TRAILING_WEEKS = 4

# --- Piso por semana individual (refinamento sobre o design original) ---
# Uma semana so entra na media se tiver pelo menos este volume. Precisa de
# pelo menos MIN_WEEKS_WITH_FLOOR das 4 semanas validas para a media contar.
WEEK_FLOOR_IMPRESSIONS = 30
MIN_WEEKS_WITH_FLOOR = 3

# --- Gatilho 1: vazamento de CTR ---
G1_MIN_IMPRESSIONS = 150   # PROVISORIO: o piso original de 500 nunca foi
# atingido por nenhuma query comercial na primeira execucao real (0 de 1642
# passaram, confirmado pelo diagnostico de funil). 150 e um chute mais
# conservador para nao ficar zerado de novo - a proxima execucao imprime a
# distribuicao real (mediana, p75, p90, p95, p99) para calibrar com dado
# em vez de outro chute.
G1_MAX_POSITION = 8
G1_CTR_ABSOLUTE_FLOOR = 0.01     # 1%
G1_CTR_DROP_RATIO = 0.40         # queda de 40% vs media

# --- Gatilho 2: queda de impressoes precoce ---
G2_TOP_N_PAGES = 20
G2_DROP_RATIO = 0.30
G2_MIN_AVG_IMPRESSIONS = 300
G2_MAX_POSITION_VARIATION = 2.0

# --- Gatilho 3: query nova emergindo ---
G3_MIN_IMPRESSIONS = 100
G3_MIN_IMPRESSIONS_QUESTION = 50
QUESTION_PATTERNS = [r"^how\b", r"^what\b", r"^why\b", r"^is\b", r"^can\b", r"^does\b",
                    r"^do\b", r"^are\b", r"^will\b", r"^should\b"]

# --- Gatilho 4: decaimento dos posts otimizados ---
G4_DROP_RATIO = 0.30
G4_MIN_AVG_CLICKS = 20
SEO_LOG_PATH = "data/seo_log_urls.csv"

# --- Gatilho 5: sinal de vida do conteudo novo ---
G5_POSITIVE_MIN_IMPRESSIONS = 50
G5_NEGATIVE_MIN_DAYS = 21
CITY_PAGES_PATH = "data/city_pages.csv"

# --- Gatilho 6: city pages radar ---
G6_NEW_PAGE_THRESHOLD = 20         # abaixo disso, pagina e tratada como nova/pequena
G6_RISE_FIRST_SIGNAL = 20         # primeira vez que bate isso = "comecou a rankear"
G6_RISE_RATIO = 0.25              # pagina estabelecida: sobe 25%+
G6_FALL_RATIO = 0.25
G6_FALL_MIN_AVG = 20
CITY_PAGE_URL_PATTERN = r"^/(?:arizona|california|nevada|florida|texas)/[a-z-]+/$"

# --- Gatilho 7: trafego de marca ---
BRAND_PATTERNS = [
    r"\bturfresh\b", r"\bturf\s*fresh\b", r"\btur\s*fresh\b", r"\bturffresh\b",
]
G7_MIN_AVG_IMPRESSIONS = 20
G7_DROP_RATIO = 0.30
G7_POSITION_DROP_ALERT = 4.0
G7_MAX_MEANINGFUL_POSITION = 20   # abaixo do top 20 a posicao e ruido, nao alerta



# ===========================================================================
# CLASSIFICADOR DE QUERY COMERCIAL/LOCAL
# RASCUNHO - Rafael precisa revisar e calibrar esta lista antes de confiar
# no Gatilho 1 de verdade. Baseado no que ja sei do ICP da TurFresh (servico
# local de limpeza de grama sintetica), nao em dado medido do site.
# ===========================================================================
COMMERCIAL_LOCAL_PATTERNS = [
    r"\bturf cleaning\b", r"\bartificial (?:turf|grass) cleaning\b",
    r"\bcleaning service\b", r"\bnear me\b", r"\bcost\b", r"\bprice\b",
    r"\bpricing\b", r"\bquote\b", r"\bhire\b", r"\bcompany\b",
    r"\bprofessional\b", r"\bbest.*service\b", r"\bhow much\b",
    # cidades prioritarias (mesma logica das 35 city pages)
    r"\bphoenix\b", r"\bpeoria\b", r"\bsanta ana\b", r"\bsan jose\b",
    r"\bthousand oaks\b", r"\bsan clemente\b", r"\bsacramento\b",
    r"\bwhittier\b", r"\bsanta clarita\b", r"\bfresno\b", r"\blos angeles\b",
    r"\briverside\b", r"\bmiami\b", r"\bjacksonville\b", r"\bfort lauderdale\b",
    r"\btampa\b", r"\bhouston\b", r"\baustin\b", r"\blas vegas\b",
    r"\birving\b", r"\bsan diego\b", r"\bmesa\b", r"\bgilbert\b",
    r"\bchandler\b", r"\btempe\b", r"\bglendale\b", r"\bnorth las vegas\b",
    r"\balhambra\b",
]
# Sinal negativo: quem pesquisa "como fazer sozinho" nao vai contratar.
DIY_PATTERNS = [
    r"\bhow to (?:clean|remove|diy)\b", r"\bmyself\b", r"\bdiy\b",
    r"\bhomemade\b", r"\bat home\b",
]


def is_commercial_local(query):
    q = query.lower()
    if any(re.search(p, q) for p in DIY_PATTERNS):
        return False
    return any(re.search(p, q) for p in COMMERCIAL_LOCAL_PATTERNS)


def is_brand(query):
    q = query.lower()
    return any(re.search(p, q) for p in BRAND_PATTERNS)


def is_question(query):
    q = query.lower().strip()
    return any(re.search(p, q) for p in QUESTION_PATTERNS)


def norm_path(url):
    p = re.sub(r"^https?://[^/]+", "", str(url)).split("?")[0].split("#")[0]
    if not p:
        p = "/"
    if len(p) > 1 and not p.endswith("/"):
        p += "/"
    return p.lower()


def load_seo_log_urls():
    """As 69 URLs vivas do SEO Log (as 44 redirecionadas ja foram excluidas
    na extracao). Retorna dict path -> {data_otimizacao, keyword, meta_title,
    h1, feito}. O 'keyword' e o Meta Title/H1 sao terreno real (o que a
    pagina de fato mira), nao mais um chute - isso deixa o Gatilho 1 e o
    Gatilho 4 checarem contra a intencao real da pagina em vez de adivinhar."""
    if not os.path.exists(SEO_LOG_PATH):
        print(f"  Aviso: {SEO_LOG_PATH} nao encontrado. Gatilho 4 fica vazio.")
        return {}
    out = {}
    with open(SEO_LOG_PATH, newline="", encoding="utf-8") as f:
        import csv
        for row in csv.DictReader(f):
            path = norm_path(row["path"])
            out[path] = {
                "data_otimizacao": row.get("data_otimizacao", ""),
                "keyword": row.get("keyword", ""),
                "meta_title": row.get("meta_title", ""),
                "h1": row.get("h1", ""),
                "feito": row.get("feito", ""),
            }
    print(f"  SEO Log: {len(out)} URLs vivas carregadas (com keyword/title/notas)")
    return out


KNOWN_ISSUE_PATTERNS = [r"\bbug\b", r"\bpending\b", r"\bneeded\b", r"\bissue\b",
                        r"\bnot showing\b", r"\bmissing\b", r"\bbroken\b"]


def extract_known_issue(feito_text):
    """
    Rafael ja documenta problemas conhecidos no campo 'Feito' quando otimiza
    um post (ex: 'Bug: H1 not showing on mobile'). Se um alerta bate numa
    pagina que ja tem isso anotado, essa e uma causa muito mais confiavel do
    que qualquer hipotese generica - e o proprio dono da pagina que registrou.
    Retorna a frase relevante, ou None se nao achar nada.
    """
    if not feito_text:
        return None
    for pattern in KNOWN_ISSUE_PATTERNS:
        m = re.search(rf"([^.]*{pattern}[^.]*)", feito_text, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return None


def keyword_overlap_ratio(query, keyword, meta_title):
    """
    Compara a query que disparou o alerta contra o Keyword Principal e o
    Meta Title reais da pagina. Overlap alto = a pagina ja mira essa
    intencao de proposito, entao 'titulo desalinhado' fica pouco provavel.
    Overlap baixo = a query pode nao ser o foco real da pagina.
    """
    def tokens(s):
        return set(w for w in re.findall(r"[a-z]+", str(s).lower())
                   if len(w) > 2 and w not in {"the","and","for","how","your",
                       "with","that","this","from","are","you"})
    q_tok = tokens(query)
    if not q_tok:
        return 0.0
    target_tok = tokens(keyword) | tokens(meta_title)
    if not target_tok:
        return 0.0
    return len(q_tok & target_tok) / len(q_tok)


def load_city_pages():
    """As 35 city pages (21 prioritarias + 14 novas). Retorna dict
    path -> {categoria, data_publicacao}."""
    if not os.path.exists(CITY_PAGES_PATH):
        print(f"  Aviso: {CITY_PAGES_PATH} nao encontrado. Gatilhos 5/6 ficam vazios.")
        return {}
    out = {}
    with open(CITY_PAGES_PATH, newline="", encoding="utf-8") as f:
        import csv
        for row in csv.DictReader(f):
            path = norm_path(row["path"])
            out[path] = {"categoria": row.get("categoria", ""),
                        "data_publicacao": row.get("data_publicacao", "")}
    print(f"  City pages: {len(out)} paginas carregadas")
    return out


# ===========================================================================
# GSC FETCH
# ===========================================================================
def get_gsc_service():
    if not GSC_CLIENT_EMAIL or not GSC_PRIVATE_KEY:
        raise RuntimeError("GSC_CLIENT_EMAIL ou GSC_PRIVATE_KEY nao definidos.")
    info = {"type": "service_account", "client_email": GSC_CLIENT_EMAIL,
            "private_key": GSC_PRIVATE_KEY,
            "token_uri": "https://oauth2.googleapis.com/token"}
    creds = service_account.Credentials.from_service_account_info(
        info, scopes=["https://www.googleapis.com/auth/webmasters.readonly"])
    return build("searchconsole", "v1", credentials=creds)


def is_garbage_query(query):
    """
    Queries que carregam uma URL dentro do texto nao representam busca
    humana real - sao ruido de bot, scraper, ou alguem colando conteudo
    errado na caixa de busca do Google. Isso importa porque um \\b (limite
    de palavra) no regex de marca trata '/' e '.' como limite, entao
    'https://turfresh.com/...' bate em '\\bturfresh\\b' mesmo nao sendo uma
    busca de marca de verdade - foi assim que uma query lixo contaminou o
    Gatilho 7. Filtrar aqui, uma vez, antes de qualquer classificador rodar,
    em vez de remendar cada regex separadamente.
    """
    q = str(query)
    if re.search(r"https?://|www\.", q, re.IGNORECASE):
        return True
    if len(q) > 100:   # buscas reais raramente passam disso
        return True
    return False


def filter_garbage_queries(df):
    if df.empty:
        return df
    mask = ~df["query"].apply(is_garbage_query)
    removed = (~mask).sum()
    if removed:
        print(f"  ({removed} queries-lixo removidas - continham URL ou eram anormalmente longas)")
    return df[mask].reset_index(drop=True)


def fetch_week(service, start_date, end_date):
    """Uma semana de dados query+page. Pagina se precisar."""
    rows_out, start_row = [], 0
    while True:
        req = {"startDate": start_date.isoformat(), "endDate": end_date.isoformat(),
               "dimensions": ["query", "page"], "rowLimit": 25000, "startRow": start_row}
        rows = service.searchanalytics().query(siteUrl=SITE_URL, body=req).execute().get("rows", [])
        if not rows:
            break
        for r in rows:
            rows_out.append({"query": r["keys"][0], "page": r["keys"][1],
                             "clicks": r["clicks"], "impressions": r["impressions"],
                             "position": r["position"]})
        if len(rows) < 25000:
            break
        start_row += 25000
    return filter_garbage_queries(pd.DataFrame(rows_out))


def fetch_trailing_weeks(service, n_weeks=N_TRAILING_WEEKS):
    """
    Retorna (semana_atual_df, [lista de N dataframes das semanas anteriores,
    mais antiga primeiro], janelas usadas).
    """
    today = date.today()
    cur_end = today - timedelta(days=WEEK_LAG_DAYS)
    cur_start = cur_end - timedelta(days=6)

    print(f"Semana atual: {cur_start} a {cur_end}")
    current = fetch_week(service, cur_start, cur_end)
    print(f"  {len(current)} linhas query+pagina\n")

    trailing = []
    windows = [(cur_start, cur_end)]
    week_start = cur_start
    for i in range(n_weeks):
        week_end = week_start - timedelta(days=1)
        week_start = week_end - timedelta(days=6)
        print(f"Semana -{i+1}: {week_start} a {week_end}")
        df = fetch_week(service, week_start, week_end)
        print(f"  {len(df)} linhas query+pagina\n")
        trailing.insert(0, df)   # mais antiga primeiro
        windows.insert(0, (week_start, week_end))

    return current, trailing, windows


# ===========================================================================
# MEDIA MOVEL COM PISO POR SEMANA
# ===========================================================================
def build_trailing_stats(trailing_weeks):
    """
    Para cada (query, page), calcula media de impressoes/clicks das semanas
    validas, exige piso MIN_WEEKS_WITH_FLOOR de MIN_WEEKS_WITH_FLOOR semanas
    com volume, e devolve tambem a posicao media (para o teste de estabilidade
    de posicao) e se a chave apareceu em QUALQUER semana anterior (Gatilho 3).

    Retorna dict: (query, page) -> {
        media_impr, media_clicks, media_position, semanas_com_piso,
        confiavel (bool), esteve_presente (bool)
    }
    """
    per_key = defaultdict(lambda: {"impr": [], "clicks": [], "pos": []})

    for week_df in trailing_weeks:
        if week_df.empty:
            continue
        for _, r in week_df.iterrows():
            k = (r["query"], r["page"])
            per_key[k]["impr"].append(float(r["impressions"]))
            per_key[k]["clicks"].append(float(r["clicks"]))
            per_key[k]["pos"].append(float(r["position"]))

    stats = {}
    for k, v in per_key.items():
        n_semanas_presentes = len(v["impr"])
        # semanas em que a chave nao apareceu contam como 0 impressoes/clicks
        impr_padded = v["impr"] + [0.0] * (N_TRAILING_WEEKS - n_semanas_presentes)
        clicks_padded = v["clicks"] + [0.0] * (N_TRAILING_WEEKS - n_semanas_presentes)

        semanas_com_piso = sum(1 for x in impr_padded if x >= WEEK_FLOOR_IMPRESSIONS)
        confiavel = semanas_com_piso >= MIN_WEEKS_WITH_FLOOR

        stats[k] = {
            "media_impr": sum(impr_padded) / N_TRAILING_WEEKS,
            "media_clicks": sum(clicks_padded) / N_TRAILING_WEEKS,
            "media_position": sum(v["pos"]) / len(v["pos"]) if v["pos"] else None,
            "semanas_com_piso": semanas_com_piso,
            "confiavel": confiavel,
            "esteve_presente": n_semanas_presentes > 0,
        }
    return stats


# ===========================================================================
# GATILHO 1 - VAZAMENTO DE CTR
# ===========================================================================
def gatilho_1_vazamento_ctr(current_df, trailing_stats, seo_log_urls=None):
    """
    Dispara quando: impressoes_semana >= 500 E posicao <= 8 E
    (CTR < 1% OU CTR caiu >= 40% vs media_4sem)

    Refinamento aplicado: a condicao de "CTR caiu vs media" so e avaliada se
    a posicao media das 4 semanas anteriores TAMBEM estava <= 8 (com folga de
    +2). Sem isso, uma pagina que acabou de subir para o top 8 essa semana
    teria uma "media" de CTR baixa so porque antes ela rankeava pior - isso
    pareceria vazamento sem ser.

    Enriquecimento com o SEO Log (quando a pagina esta nas 67 vivas): em vez
    de so hipotetizar "talvez o titulo nao combine", compara a query real
    contra o Keyword Principal e o Meta Title que Rafael de fato escreveu
    para aquela pagina. Se baterem bem, a hipotese de titulo desalinhado
    perde forca de verdade, nao so por suposicao. E se o campo "Feito" ja
    tem um bug ou pendencia documentada, isso e mostrado direto - e mais
    confiavel que qualquer hipotese gerada.
    """
    alerts = []
    if current_df.empty:
        return alerts
    seo_log_urls = seo_log_urls or {}

    for _, r in current_df.iterrows():
        query, page = r["query"], r["page"]
        page_is_tracked = norm_path(page) in seo_log_urls
        if not is_commercial_local(query) and not page_is_tracked:
            continue

        impr = float(r["impressions"])
        pos = float(r["position"])
        clicks = float(r["clicks"])

        if impr < G1_MIN_IMPRESSIONS or pos > G1_MAX_POSITION:
            continue

        ctr = clicks / impr if impr else 0
        k = (query, page)
        st = trailing_stats.get(k)

        by_absolute = ctr < G1_CTR_ABSOLUTE_FLOOR

        by_drop = False
        drop_pct = None
        if st and st["confiavel"] and st["media_clicks"] > 0 and st["media_impr"] > 0:
            media_ctr = st["media_clicks"] / st["media_impr"]
            # so avalia queda se a posicao ja estava boa nas semanas anteriores
            posicao_ja_estavel = (st["media_position"] is not None
                                  and st["media_position"] <= G1_MAX_POSITION + 2)
            if media_ctr > 0 and posicao_ja_estavel:
                drop_pct = (media_ctr - ctr) / media_ctr
                by_drop = drop_pct >= G1_CTR_DROP_RATIO

        if not (by_absolute or by_drop):
            continue

        if by_absolute and by_drop:
            motivo = f"Absolute CTR of {ctr*100:.2f}% (below the 1% floor) AND dropped {drop_pct*100:.0f}% vs average"
        elif by_absolute:
            motivo = f"CTR of {ctr*100:.2f}%, below the 1% absolute floor for position {pos:.1f}"
        else:
            motivo = f"CTR dropped {drop_pct*100:.0f}% vs the previous 4-week average"

        confianca = "HIGH" if (st and st["confiavel"]) else "LOW (limited history)"

        # --- Zero estrutural: CTR literalmente 0% (nao so "abaixo de 1%") numa
        # posicao boa (top 8) com volume real. Isso e um padrao diferente de
        # "titulo fraco" - e a assinatura classica de uma citacao de AI
        # Overview ou outro recurso que mostra a pagina como fonte sem gerar
        # clique nenhum. Nao ha correcao de copy para isso, entao nao deve
        # competir por atencao como se fosse um problema resolvivel.
        is_structural_zero = ctr < 0.001

        # --- enriquecimento com dado real da pagina, quando disponivel ---
        seo_log_hyp = None
        seo_log_verify = None
        seo_log_action = None
        actionable = True   # vira tarefa por padrao; alguns casos abaixo desligam isso
        page_path = norm_path(page)
        meta = seo_log_urls.get(page_path)

        if is_structural_zero:
            seo_log_hyp = (f"CTR is literally 0% at position {pos:.1f} with real volume "
                           f"({int(impr)} impressions) - not just low. This is the typical "
                           f"signature of an AI Overview or other feature citing this page as "
                           f"a source without sending a click, not a weak title.")
            seo_log_verify = (f"Search this exact query and check whether an AI Overview or "
                              f"other feature is citing this page instead of showing it as a "
                              f"normal clickable result.")
            seo_log_action = "No copy fix applies here. If confirmed as an AI Overview citation, this is worth knowing but not worth a task - track it, don't act on it."
            actionable = False
        elif meta:
            known_issue = extract_known_issue(meta.get("feito", ""))
            overlap = keyword_overlap_ratio(query, meta.get("keyword", ""), meta.get("meta_title", ""))

            if known_issue:
                seo_log_hyp = (f"This page already has a documented issue from its last "
                               f"optimization: \"{known_issue}\"")
                seo_log_verify = "Confirm whether this documented issue is still unresolved."
                seo_log_action = "Fix the documented issue - this is more likely the cause than a title/meta mismatch."
            elif overlap >= 0.5:
                seo_log_hyp = (f"This page is already targeted at \"{meta.get('keyword','')}\" "
                               f"(meta title: \"{meta.get('meta_title','')}\"), which closely "
                               f"matches this query - a title/meta rewrite is unlikely to be the fix.")
                seo_log_verify = ("Check the live SERP for a map pack, People Also Ask, or AI "
                                  "Overview above position " + f"{pos:.0f} - that is the more "
                                  "likely cause given the title already matches.")
                seo_log_action = "No title/meta change to make here. If a SERP feature is confirmed, there is usually no direct fix - not worth a task."
                actionable = False
            else:
                seo_log_hyp = (f"This page is logged as targeting \"{meta.get('keyword','')}\" "
                               f"(meta title: \"{meta.get('meta_title','')}\"), which does not "
                               f"closely match this triggering query - the page may be ranking "
                               f"for this query almost by accident.")
                seo_log_verify = ("Decide whether this query is worth targeting on this page at "
                                  "all, or whether a dedicated page/section would serve it better.")
                seo_log_action = "Only adjust this page's title/meta if you decide this query belongs to it. Otherwise, this may not be worth optimizing for."

        alert = {
            "gatilho": "1. CTR Gap",
            "query": query,
            "pagina": page,
            "posicao": round(pos, 1),
            "impressoes_semana": int(impr),
            "clicks_semana": int(clicks),
            "ctr_semana": f"{ctr*100:.2f}%",
            "media_ctr_4sem": (f"{(st['media_clicks']/st['media_impr'])*100:.2f}%"
                               if st and st["media_impr"] > 0 else "no history"),
            "motivo": motivo,
            "confianca": confianca,
            "actionable": actionable,
        }
        if seo_log_hyp:
            alert["hypothesis_override"] = seo_log_hyp
            alert["verify_override"] = seo_log_verify
            alert["action_override"] = seo_log_action
        alerts.append(alert)

    alerts.sort(key=lambda x: -x["impressoes_semana"])
    return alerts


# ===========================================================================
# GATILHO 2 - QUEDA DE IMPRESSOES PRECOCE
# ===========================================================================
def gatilho_2_queda_precoce(current_df, trailing_stats):
    """
    Top 20 paginas + queries comerciais. Dispara quando impressoes cairam
    >=30% vs media_4sem (media confiavel, piso 300/sem) E posicao ficou
    estavel (variacao < 2). Posicao estavel + impressao caindo = Google
    mostrando menos, nao voce piorando (AI Overview, perda de feature).
    """
    alerts = []
    if current_df.empty:
        return alerts

    page_impr = current_df.groupby("page")["impressions"].sum().sort_values(ascending=False)
    top_pages = set(page_impr.head(G2_TOP_N_PAGES).index)
    money_queries = set(current_df[current_df["query"].apply(is_commercial_local)]["query"])

    for _, r in current_df.iterrows():
        query, page = r["query"], r["page"]
        if page not in top_pages and query not in money_queries:
            continue

        k = (query, page)
        st = trailing_stats.get(k)
        if not st or not st["confiavel"] or st["media_impr"] < G2_MIN_AVG_IMPRESSIONS:
            continue

        impr = float(r["impressions"])
        pos = float(r["position"])
        drop = (st["media_impr"] - impr) / st["media_impr"]
        if drop < G2_DROP_RATIO:
            continue

        pos_variation = abs(pos - st["media_position"]) if st["media_position"] else 999
        if pos_variation >= G2_MAX_POSITION_VARIATION:
            continue   # posicao tambem mudou - nao e o padrao "Google mostrando menos"

        alerts.append({
            "gatilho": "2. Early Impression Drop",
            "query": query, "pagina": page, "posicao": round(pos, 1),
            "posicao_media_4sem": round(st["media_position"], 1),
            "impressoes_semana": int(impr),
            "media_impr_4sem": round(st["media_impr"], 0),
            "queda_pct": f"{drop*100:.0f}%",
            "motivo": (f"Impressions dropped {drop*100:.0f}% vs average (position stayed "
                      f"stable: {pos:.1f} vs average {st['media_position']:.1f}). "
                      f"Google is showing this page/query less, not a "
                      f"ranking problem."),
            "confianca": "HIGH",
        })

    alerts.sort(key=lambda x: -x["impressoes_semana"])
    return alerts


# ===========================================================================
# GATILHO 3 - QUERY NOVA EMERGINDO
# ===========================================================================
def gatilho_3_query_nova(current_df, trailing_stats):
    """
    Query ausente nas 4 semanas anteriores E impressoes_semana >= 100.
    Piso cai para 50 se for pergunta (candidata a calendario de AI).
    """
    alerts = []
    if current_df.empty:
        return alerts

    seen_queries = current_df.groupby("query")["impressions"].sum()
    for query, impr in seen_queries.items():
        # ja apareceu em alguma semana anterior para QUALQUER pagina?
        already_seen = any(st["esteve_presente"] for k, st in trailing_stats.items()
                           if k[0] == query)
        if already_seen:
            continue

        pergunta = is_question(query)
        piso = G3_MIN_IMPRESSIONS_QUESTION if pergunta else G3_MIN_IMPRESSIONS
        if impr < piso:
            continue

        page = current_df[current_df["query"] == query].iloc[0]["page"]
        pos = current_df[current_df["query"] == query].iloc[0]["position"]

        alerts.append({
            "gatilho": "3. Emerging New Query",
            "query": query, "pagina": page, "posicao": round(float(pos), 1),
            "impressoes_semana": int(impr),
            "tipo": "QUESTION - AI content calendar candidate" if pergunta else "standard",
            "motivo": (f"Query did not exist in the previous 4 weeks. "
                      f"{'It is a question, a good candidate for AI Overview content.' if pergunta else ''}"),
            "confianca": "HIGH" if impr >= 200 else "MEDIUM",
        })

    alerts.sort(key=lambda x: -x["impressoes_semana"])
    return alerts


# ===========================================================================
# GATILHO 4 - DECAIMENTO DOS POSTS OTIMIZADOS
# ===========================================================================
def gatilho_4_decaimento_posts(current_df, trailing_stats, seo_log_urls):
    """
    Escopo: as 69 URLs vivas do SEO Log (as 44 redirecionadas ja foram
    excluidas na extracao dos dados). Clicks cairam >=30% vs media_4sem,
    com piso de 20 clicks/sem para nao alertar ruido de posts pequenos.
    """
    alerts = []
    if current_df.empty or not seo_log_urls:
        return alerts

    current_df = current_df.copy()
    current_df["path"] = current_df["page"].apply(norm_path)
    scoped = current_df[current_df["path"].isin(seo_log_urls.keys())]

    page_clicks = scoped.groupby("path")["clicks"].sum()
    page_impr = scoped.groupby("path")["impressions"].sum()
    page_pos = scoped.groupby("path")["position"].mean()

    # agrega trailing_stats por pagina (nao por query+pagina)
    page_trailing = defaultdict(lambda: {"clicks": [0.0]*N_TRAILING_WEEKS})
    for (query, page), st in trailing_stats.items():
        path = norm_path(page)
        if path in seo_log_urls:
            # aproximacao: soma media_clicks de todas as queries dessa pagina
            page_trailing[path]["clicks"] = [
                page_trailing[path]["clicks"][0] + st["media_clicks"]]

    for path in page_clicks.index:
        media_clicks = page_trailing.get(path, {}).get("clicks", [0])[0]
        if media_clicks < G4_MIN_AVG_CLICKS:
            continue
        clicks_now = float(page_clicks[path])
        drop = (media_clicks - clicks_now) / media_clicks if media_clicks else 0
        if drop < G4_DROP_RATIO:
            continue

        meta = seo_log_urls.get(path, {})
        known_issue = extract_known_issue(meta.get("feito", ""))

        alerts.append({
            "gatilho": "4. Optimized Post Decay",
            "pagina": path,
            "clicks_semana": int(clicks_now),
            "media_clicks_4sem": round(media_clicks, 1),
            "impressoes_semana": int(page_impr.get(path, 0)),
            "posicao": round(float(page_pos.get(path, 0)), 1),
            "queda_pct": f"{drop*100:.0f}%",
            "data_otimizacao": meta.get("data_otimizacao", ""),
            "keyword": meta.get("keyword", ""),
            "known_issue": known_issue,
            "motivo": f"Clicks dropped {drop*100:.0f}% vs the previous 4-week average.",
            "confianca": "HIGH" if media_clicks >= 40 else "MEDIUM",
        })

    alerts.sort(key=lambda x: -x["media_clicks_4sem"])
    return alerts


# ===========================================================================
# GATILHO 5 - SINAL DE VIDA DO CONTEUDO NOVO
# ===========================================================================
def gatilho_5_sinal_vida(current_df, city_pages, run_date):
    """
    Escopo: city pages com data de publicacao conhecida, publicadas nos
    ultimos ~35 dias (folga sobre os 30 para nao perder o corte por causa
    do lag de dados do GSC).
    Positivo: primeira semana com impressoes >= 50.
    Negativo: publicada ha >= 21 dias E impressoes = 0 (problema de indexacao).
    """
    alerts = []
    if current_df.empty or not city_pages:
        return alerts

    current_df = current_df.copy()
    current_df["path"] = current_df["page"].apply(norm_path)
    page_impr = current_df.groupby("path")["impressions"].sum()

    for path, meta in city_pages.items():
        data_pub_str = meta.get("data_publicacao", "")
        if not data_pub_str or "verificar" in data_pub_str.lower():
            continue   # sem data confiavel, nao da para avaliar "sinal de vida"

        dt = _parse_date_flex(data_pub_str)
        if dt is None:
            continue

        dias_desde_publicacao = (run_date - dt).days
        if dias_desde_publicacao < 0 or dias_desde_publicacao > 35:
            continue

        impr = float(page_impr.get(path, 0))

        if dias_desde_publicacao <= 7 and impr >= G5_POSITIVE_MIN_IMPRESSIONS:
            alerts.append({
                "gatilho": "5. Content Life Signal (Positive)",
                "pagina": path, "dias_desde_publicacao": dias_desde_publicacao,
                "impressoes_semana": int(impr),
                "motivo": f"First week with {int(impr)} impressions - Google has started testing the page.",
                "confianca": "HIGH",
            })
        elif dias_desde_publicacao >= G5_NEGATIVE_MIN_DAYS and impr == 0:
            alerts.append({
                "gatilho": "5. Content Life Signal (Negative)",
                "pagina": path, "dias_desde_publicacao": dias_desde_publicacao,
                "impressoes_semana": 0,
                "motivo": (f"Published {dias_desde_publicacao} days ago, ZERO impressions. "
                          f"Possible indexing issue - check in GSC (URL Inspection)."),
                "confianca": "HIGH",
            })

    return alerts


def _parse_date_flex(s):
    """Datas no SEO Log vem em formatos inconsistentes (07 Mai, 10 Jul 2026,
    Verificar). Tenta alguns formatos comuns em portugues; devolve None se
    nao conseguir, para o chamador decidir pular em vez de quebrar."""
    s = str(s).strip()
    meses = {"jan":1,"fev":2,"mar":3,"abr":4,"mai":5,"jun":6,"jul":7,"ago":8,
              "set":9,"out":10,"nov":11,"dez":12}
    m = re.match(r"(\d{1,2})\s+([a-zA-Z]{3})\s*(\d{4})?", s)
    if not m:
        return None
    dia, mes_str, ano = m.groups()
    mes = meses.get(mes_str.lower()[:3])
    if not mes:
        return None
    ano = int(ano) if ano else date.today().year
    try:
        return date(ano, mes, int(dia))
    except ValueError:
        return None


# ===========================================================================
# GATILHO 6 - CITY PAGES (SUBINDO / CAINDO)
# ===========================================================================
def gatilho_6_city_pages(current_df, trailing_stats, city_pages):
    """
    Escopo: as 35 city pages cadastradas + qualquer URL que bata no padrao
    /{estado}/{cidade}/ (deteccao automatica, cresce sozinha com paginas novas).
    """
    alerts = []
    if current_df.empty:
        return alerts

    current_df = current_df.copy()
    current_df["path"] = current_df["page"].apply(norm_path)

    is_city = (current_df["path"].isin(city_pages.keys())
              | current_df["path"].str.match(CITY_PAGE_URL_PATTERN))
    scoped = current_df[is_city]

    page_impr_now = scoped.groupby("path")["impressions"].sum()
    page_clicks_now = scoped.groupby("path")["clicks"].sum()

    page_trailing = defaultdict(lambda: [0.0] * N_TRAILING_WEEKS)
    for (query, page), st in trailing_stats.items():
        path = norm_path(page)
        if path in page_impr_now.index:
            # aproxima somando as medias de impressao de todas as queries da pagina
            pass  # tratado abaixo de forma agregada

    # media por pagina: soma media_impr de todas as (query,page) daquela pagina
    page_media = defaultdict(float)
    for (query, page), st in trailing_stats.items():
        path = norm_path(page)
        if path in page_impr_now.index:
            page_media[path] += st["media_impr"]

    for path in page_impr_now.index:
        impr_now = float(page_impr_now[path])
        media = page_media.get(path, 0.0)

        if media < G6_NEW_PAGE_THRESHOLD:
            if impr_now >= G6_RISE_FIRST_SIGNAL:
                alerts.append({
                    "gatilho": "6. City Page Rising",
                    "pagina": path, "impressoes_semana": int(impr_now),
                    "media_impr_4sem": round(media, 0),
                    "motivo": f"New/small page has started ranking: {int(impr_now)} impressions this week.",
                    "confianca": "MEDIUM",
                })
        else:
            variacao = (impr_now - media) / media
            if variacao >= G6_RISE_RATIO:
                alerts.append({
                    "gatilho": "6. City Page Rising",
                    "pagina": path, "impressoes_semana": int(impr_now),
                    "media_impr_4sem": round(media, 0),
                    "motivo": f"Impressions rose {variacao*100:.0f}% vs the 4-week average.",
                    "confianca": "HIGH",
                })
            elif -variacao >= G6_FALL_RATIO and media >= G6_FALL_MIN_AVG:
                alerts.append({
                    "gatilho": "6. City Page Falling",
                    "pagina": path, "impressoes_semana": int(impr_now),
                    "media_impr_4sem": round(media, 0),
                    "motivo": f"Impressions dropped {-variacao*100:.0f}% vs the 4-week average.",
                    "confianca": "HIGH",
                })

    alerts.sort(key=lambda x: -x["impressoes_semana"])
    return alerts


# ===========================================================================
# GATILHO 7 - TRAFEGO DE MARCA
# ===========================================================================
def gatilho_7_marca(current_df, trailing_stats):
    """
    Duas formas de disparar, porque marca tem um risco que query comum nao
    tem - alguem pode superar voce no seu proprio nome:
      A) impressoes/clicks caindo com posicao estavel (Google mostrando
         menos a marca - AI Overview, Knowledge Panel, etc.)
      B) POSICAO da marca caindo (concorrente/terceiro ultrapassando voce
         na busca do seu proprio nome - prioridade maxima, mesmo com volume
         baixo, porque isso e estrutural, nao ruido).
    """
    alerts = []
    if current_df.empty:
        return alerts

    brand_current = current_df[current_df["query"].apply(is_brand)]
    if brand_current.empty:
        return alerts

    for _, r in brand_current.iterrows():
        query, page = r["query"], r["page"]
        k = (query, page)
        st = trailing_stats.get(k)
        if not st:
            continue

        impr = float(r["impressions"])
        pos = float(r["position"])

        # B) posicao caindo - grave, mas so se a confianca for real. Sem
        # isso, qualquer variante de marca de cauda longa com 2-3 impressoes
        # (posicao naturalmente ruidosa) disparava "urgente" por oscilacao
        # normal, e query em posicao 90+ (invisivel) alertava mesmo sem
        # significado operacional nenhum.
        if (st["confiavel"] and st["media_position"] is not None
                and st["media_position"] <= G7_MAX_MEANINGFUL_POSITION):
            pos_drop = pos - st["media_position"]
            if pos_drop >= G7_POSITION_DROP_ALERT:
                alerts.append({
                    "gatilho": "7. Brand - POSITION DROPPING",
                    "query": query, "pagina": page, "posicao": round(pos, 1),
                    "posicao_media_4sem": round(st["media_position"], 1),
                    "impressoes_semana": int(impr),
                    "motivo": (f"Brand position dropped from {st['media_position']:.1f} to "
                              f"{pos:.1f}. Someone may be outranking you on your own "
                              f"name - check manually as a priority."),
                    "confianca": "HIGH",
                    "urgencia": "MAXIMUM",
                })
                continue   # no need to also check the other trigger

        # A) volume caindo, posicao estavel
        if not st["confiavel"] or st["media_impr"] < G7_MIN_AVG_IMPRESSIONS:
            continue
        drop = (st["media_impr"] - impr) / st["media_impr"]
        if drop < G7_DROP_RATIO:
            continue
        pos_variation = abs(pos - st["media_position"]) if st["media_position"] else 999
        if pos_variation >= G2_MAX_POSITION_VARIATION:
            continue

        alerts.append({
            "gatilho": "7. Brand - Traffic Dropping",
            "query": query, "pagina": page, "posicao": round(pos, 1),
            "impressoes_semana": int(impr), "media_impr_4sem": round(st["media_impr"], 0),
            "motivo": (f"Brand impressions dropped {drop*100:.0f}% vs average, position "
                      f"stable. Google is showing the brand less - check Knowledge Panel, "
                      f"GBP, or an AI Overview appearing in your site's place."),
            "confianca": "HIGH",
            "urgencia": "HIGH",
        })

    alerts.sort(key=lambda x: (x.get("urgencia") != "MAXIMUM", -x["impressoes_semana"]))
    return alerts


# ===========================================================================
# GATILHO 9 - CANIBALIZACAO
# Nao existia antes. Achado real: "artificial grass cleaning san jose"
# rankeava em /california/san-jose/ (pos 1.7) E /california/bay-area/
# (pos 7.7) ao mesmo tempo - G1 e G2 tratavam isso como dois problemas
# separados quando e um problema estrutural so: duas paginas brigando
# pela mesma query.
# ===========================================================================
G9_MIN_IMPRESSIONS = 50
G9_MAX_POSITION = 30


def gatilho_9_canibalizacao(current_df, dead_set=None):
    """
    Para cada query, se 2+ paginas distintas rankeiam com volume real, isso
    e canibalizacao - o Google nao sabe qual pagina sua deveria vencer, e
    isso normalmente arrasta a posicao da melhor pagina pra baixo tambem.
    """
    alerts = []
    if current_df.empty:
        return alerts
    dead_set = dead_set or set()

    for query, grp in current_df.groupby("query"):
        if query in dead_set:
            continue
        competing = grp[grp["impressions"] >= G9_MIN_IMPRESSIONS]
        competing = competing[competing["position"] <= G9_MAX_POSITION]
        pages = competing["page"].unique()
        if len(pages) < 2:
            continue

        competing = competing.sort_values("position")
        best = competing.iloc[0]
        others = competing.iloc[1:]

        alerts.append({
            "gatilho": "9. Cannibalization",
            "query": query,
            "pagina": best["page"],
            "posicao": round(float(best["position"]), 1),
            "impressoes_semana": int(competing["impressions"].sum()),
            "n_paginas": len(pages),
            "outras_paginas": " | ".join(f"{norm_path(p)} (pos {pos:.1f})"
                                         for p, pos in zip(others["page"], others["position"])),
            "motivo": (f"{len(pages)} of your own pages compete for this query. Best position "
                      f"is {best['position']:.1f} on {norm_path(best['page'])}, but the other "
                      f"{len(pages)-1} page(s) split impressions and likely hold the best page back."),
            "confianca": "HIGH",
        })

    alerts.sort(key=lambda x: -x["impressoes_semana"])
    return alerts


# ===========================================================================
# DETECTOR DE PADRAO SISTEMICO
# Achado real: 8 city pages diferentes cairam entre 66% e 83% de impressao
# na MESMA semana, todas com posicao estavel - muito estreito pra serem 8
# problemas independentes. Isso junta esses casos numa descoberta so, em
# vez de espalhar como N tarefas identicas.
# ===========================================================================
SYSTEMIC_MIN_PAGES = 3
SYSTEMIC_DROP_BAND = 20   # pontos percentuais de largura da faixa considerada "o mesmo evento"


def detect_systemic_pattern(g2, g6):
    """
    Junta alertas de G2 (queda precoce) e G6 (city page caindo) que
    compartilham uma faixa de queda muito estreita, em paginas distintas,
    na mesma semana - assinatura de um evento so (ex: uma mudanca de SERP
    que afeta uma familia inteira de queries) em vez de N problemas
    independentes. So agrupa se houver evidencia real de cluster; nao forca
    juncao em alertas que nao se parecem.
    """
    candidates = []
    for a in g2:
        try:
            drop = float(str(a.get("queda_pct", "0")).replace("%", ""))
        except ValueError:
            continue
        candidates.append({"page": norm_path(a["pagina"]), "drop": drop,
                           "impr": a.get("impressoes_semana", 0), "source": "G2",
                           "query": a.get("query", "")})
    for a in g6:
        if a.get("gatilho") != "6. City Page Falling":
            continue
        media = a.get("media_impr_4sem", 0) or 0
        impr = a.get("impressoes_semana", 0) or 0
        if media <= 0:
            continue
        drop = (media - impr) / media * 100
        candidates.append({"page": norm_path(a["pagina"]), "drop": drop,
                           "impr": impr, "source": "G6", "query": ""})

    if len(candidates) < SYSTEMIC_MIN_PAGES:
        return None

    candidates.sort(key=lambda c: c["drop"])
    best_cluster = []
    for c in candidates:
        band = [x for x in candidates if abs(x["drop"] - c["drop"]) <= SYSTEMIC_DROP_BAND / 2]
        distinct_pages = {x["page"] for x in band}
        if len(distinct_pages) > len({x["page"] for x in best_cluster}):
            best_cluster = band

    distinct_pages = {c["page"] for c in best_cluster}
    if len(distinct_pages) < SYSTEMIC_MIN_PAGES:
        return None

    drops = [c["drop"] for c in best_cluster]
    return {
        "gatilho": "SYSTEMIC. Site-wide visibility drop",
        "n_pages": len(distinct_pages),
        "pages": sorted(distinct_pages),
        "drop_range": f"{min(drops):.0f}%-{max(drops):.0f}%",
        "total_impressions": sum(c["impr"] for c in best_cluster),
        "motivo": (f"{len(distinct_pages)} different pages dropped {min(drops):.0f}%-"
                  f"{max(drops):.0f}% in impressions this week, all with stable position. "
                  f"That range is too tight to be independent problems - this is very "
                  f"likely one event (a new AI Overview or SERP feature rollout affecting "
                  f"this query family) hitting all of them at once."),
    }


# RADAR - VISIBILIDADE COMPLETA (nao dispara alerta, so mostra o estado)
# ===========================================================================
def build_radar(current_df, trailing_stats, all_alerts):
    """
    Uma linha por pagina com volume real: impressoes/clicks atuais, media de
    4 semanas, tendencia, e se algum gatilho pegou ela essa semana. Isso
    complementa os 7 gatilhos - eles avisam quando algo precisa de acao, o
    Radar deixa ver o estado de tudo, mesmo o que nao disparou nada.
    """
    if current_df.empty:
        return []

    flagged_pages = defaultdict(list)
    for a in all_alerts:
        if "pagina" in a:
            flagged_pages[norm_path(a["pagina"])].append(a["gatilho"])

    page_agg = current_df.groupby("page", as_index=False).agg(
        impressoes=("impressions", "sum"), clicks=("clicks", "sum"),
        posicao=("position", "mean"))
    page_agg = page_agg[page_agg["impressoes"] >= 30]   # corta ruido de cauda longa

    page_media = defaultdict(float)
    page_media_clicks = defaultdict(float)
    for (query, page), st in trailing_stats.items():
        path = norm_path(page)
        page_media[path] += st["media_impr"]
        page_media_clicks[path] += st["media_clicks"]

    rows = []
    for _, r in page_agg.iterrows():
        path = norm_path(r["page"])
        media = page_media.get(path, 0)
        tendencia = ""
        if media > 0:
            var = (r["impressoes"] - media) / media
            tendencia = f"{var*100:+.0f}%"

        rows.append({
            "pagina": path,
            "impressoes_semana": int(r["impressoes"]),
            "clicks_semana": int(r["clicks"]),
            "posicao": round(float(r["posicao"]), 1),
            "media_impr_4sem": round(media, 0),
            "tendencia": tendencia,
            "gatilhos_ativos": ", ".join(sorted(set(flagged_pages.get(path, [])))) or "-",
        })

    rows.sort(key=lambda x: -x["impressoes_semana"])
    return rows


# ===========================================================================
# DIAGNOSTICO
# ===========================================================================
def print_funnel_diagnostic(current_df):
    """
    0 alertas pode ser 'esta tudo bem' ou pode ser 'o filtro nunca chega no
    final'. Sem isso os dois casos ficam indistinguiveis - e nao vou
    reportar saude do site sem checar qual dos dois esta acontecendo.
    Roda toda semana, nao so uma vez - se o classificador ficar ruim com o
    tempo (site muda), isso aparece aqui antes de virar zero alertas silencioso.
    """
    print("=" * 70)
    print("DIAGNOSTICO DO FUNIL")
    print("=" * 70)
    total_queries = current_df["query"].nunique()
    comm = current_df[current_df["query"].apply(is_commercial_local)]
    n_comm = comm["query"].nunique()
    n_comm_impr = comm[comm["impressions"] >= G1_MIN_IMPRESSIONS]["query"].nunique()
    n_comm_impr_pos = comm[(comm["impressions"] >= G1_MIN_IMPRESSIONS)
                           & (comm["position"] <= G1_MAX_POSITION)]["query"].nunique()
    brand = current_df[current_df["query"].apply(is_brand)]
    print(f"  Queries unicas na semana: {total_queries}")
    print(f"  Classificadas comercial/local: {n_comm}")
    print(f"    ...com {G1_MIN_IMPRESSIONS}+ impr numa semana: {n_comm_impr}")
    print(f"    ...E posicao <= {G1_MAX_POSITION}: {n_comm_impr_pos}")
    print(f"  Classificadas como marca: {brand['query'].nunique()}")
    if n_comm == 0:
        print("\n  ALERTA: regex comercial/local nao capturou nenhuma query.")
    if brand.empty:
        print("\n  ALERTA: regex de marca nao capturou nenhuma query - confirmar variacoes do nome.")

    # Distribuicao real de impressoes entre queries comerciais - para
    # calibrar G1_MIN_IMPRESSIONS com evidencia, nao com outro chute.
    if n_comm > 0:
        impr_por_query = comm.groupby("query")["impressions"].sum()
        pcts = impr_por_query.quantile([0.5, 0.75, 0.90, 0.95, 0.99])
        print(f"\n  Distribuicao de impressoes/semana entre as {n_comm} queries comerciais:")
        print(f"    mediana: {pcts[0.5]:.0f}  |  p75: {pcts[0.75]:.0f}  |  "
              f"p90: {pcts[0.90]:.0f}  |  p95: {pcts[0.95]:.0f}  |  p99: {pcts[0.99]:.0f}")
        if n_comm_impr == 0 and pcts[0.99] < G1_MIN_IMPRESSIONS:
            print(f"    O piso de {G1_MIN_IMPRESSIONS} esta acima do p99 - nenhuma query")
            print(f"    individual vai bater isso numa semana so. Considerar baixar para")
            print(f"    perto do p90 ({pcts[0.90]:.0f}) na proxima calibracao.")
    print()


# ===========================================================================
# EXCEL
# ===========================================================================
HEADER = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def _sheet(wb, title, rows, columns):
    ws = wb.create_sheet(title)
    for i, (key, h, w) in enumerate(columns, 1):
        c = ws.cell(1, i, h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = HEADER
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26
    for ri, row in enumerate(rows, 2):
        for ci, (key, _, _) in enumerate(columns, 1):
            cell = ws.cell(ri, ci, row.get(key, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if key == "urgencia" and row.get("urgencia") == "MAXIMUM":
                cell.fill = RED
                cell.font = Font(bold=True)
            elif key == "confianca":
                fill = {"HIGH": GREEN, "MEDIUM": YELLOW, "LOW": ORANGE}.get(row.get("confianca"))
                if fill:
                    cell.fill = fill
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows)+1}"
    return ws


def write_report(g1, g2, g3, g4, g5, g6, g7, g8, g9, systemic, radar, run_date, path="turfresh_alertas.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"TurFresh - Weekly Alerts - {run_date.isoformat()}"
    ws["A1"].font = Font(bold=True, size=14)
    row_n = 3
    if systemic:
        ws[f"A{row_n}"] = (f"SYSTEMIC PATTERN: {systemic['n_pages']} pages dropped "
                           f"{systemic['drop_range']} in impressions this week - see below.")
        ws[f"A{row_n}"].font = Font(bold=True, color="D32F2F")
        row_n += 1
    ws[f"A{row_n}"] = f"Trigger 1 (CTR gap): {len(g1)}"; row_n += 1
    ws[f"A{row_n}"] = f"Trigger 2 (early impression drop): {len(g2)}"; row_n += 1
    ws[f"A{row_n}"] = f"Trigger 3 (emerging query): {len(g3)}"; row_n += 1
    ws[f"A{row_n}"] = f"Trigger 4 (optimized post decay): {len(g4)}"; row_n += 1
    ws[f"A{row_n}"] = f"Trigger 5 (content life signal): {len(g5)}"; row_n += 1
    ws[f"A{row_n}"] = f"Trigger 6 (city pages): {len(g6)}"; row_n += 1
    ws[f"A{row_n}"] = f"Trigger 7 (brand): {len(g7)}"; row_n += 1
    ws[f"A{row_n}"] = f"Trigger 8 (blog CTA): {len(g8)} campaigns with a session this week"; row_n += 1
    ws[f"A{row_n}"] = f"Trigger 9 (cannibalization): {len(g9)}"; row_n += 1
    ws.column_dimensions["A"].width = 55

    G1_COLS = [("query","Query",40),("pagina","Page",38),("posicao","Pos",8),
               ("impressoes_semana","Impr/wk",10),("ctr_semana","CTR",9),
               ("media_ctr_4sem","4wk avg CTR",13),("confianca","Confidence",11),
               ("motivo","Reason",60)]
    _sheet(wb, "G1 CTR Gap", g1, G1_COLS)

    G2_COLS = [("query","Query",40),("pagina","Page",38),("posicao","Pos",8),
               ("posicao_media_4sem","4wk avg pos",13),
               ("impressoes_semana","Impr/wk",10),("media_impr_4sem","4wk avg",12),
               ("queda_pct","Drop",9),("confianca","Confidence",11),("motivo","Reason",60)]
    _sheet(wb, "G2 Early Drop", g2, G2_COLS)

    G3_COLS = [("query","Query",40),("pagina","Page",38),("posicao","Pos",8),
               ("impressoes_semana","Impr/wk",10),("tipo","Type",28),
               ("confianca","Confidence",11),("motivo","Reason",55)]
    _sheet(wb, "G3 New Query", g3, G3_COLS)

    G4_COLS = [("pagina","Page",42),("clicks_semana","Clicks/wk",11),
               ("media_clicks_4sem","4wk avg",12),("impressoes_semana","Impr/wk",10),
               ("posicao","Pos",8),("queda_pct","Drop",9),
               ("data_otimizacao","Optimized on",11),("confianca","Confidence",11),
               ("motivo","Reason",55)]
    _sheet(wb, "G4 Post Decay", g4, G4_COLS)

    G5_COLS = [("pagina","Page",42),("dias_desde_publicacao","Days",8),
               ("impressoes_semana","Impr/wk",10),("confianca","Confidence",11),
               ("motivo","Reason",65)]
    _sheet(wb, "G5 Life Signal", g5, G5_COLS)

    G6_COLS = [("gatilho","Direction",22),("pagina","Page",42),
               ("impressoes_semana","Impr/wk",10),("media_impr_4sem","4wk avg",12),
               ("confianca","Confidence",11),("motivo","Reason",60)]
    _sheet(wb, "G6 City Pages", g6, G6_COLS)

    G7_COLS = [("urgencia","Urgency",10),("gatilho","Type",25),("query","Query",30),
               ("pagina","Page",38),("posicao","Pos",8),
               ("posicao_media_4sem","4wk avg pos",13),
               ("impressoes_semana","Impr/wk",10),("confianca","Confidence",11),
               ("motivo","Reason",60)]
    _sheet(wb, "G7 Brand", g7, G7_COLS)

    G8_COLS = [("campanha","Campaign (post)",42),("sessoes_semana","Sessions this week",16),
               ("sessoes_semana_anterior","Sessions previous wk",18),
               ("variacao","Change",12),("alerta","Alert",55)]
    _sheet(wb, "G8 Blog CTA", g8, G8_COLS)

    G9_COLS = [("query","Query",38),("pagina","Best page",38),("posicao","Pos",8),
               ("n_paginas","# pages",9),("impressoes_semana","Impr/wk",10),
               ("outras_paginas","Competing pages",55),("confianca","Confidence",11),
               ("motivo","Reason",60)]
    _sheet(wb, "G9 Cannibalization", g9, G9_COLS)

    RADAR_COLS = [("pagina","Page",42),("impressoes_semana","Impr/wk",11),
                  ("clicks_semana","Clicks/wk",11),("posicao","Pos",8),
                  ("media_impr_4sem","4wk avg",12),("tendencia","Trend",11),
                  ("gatilhos_ativos","Active triggers",35)]
    _sheet(wb, "Radar", radar, RADAR_COLS)

    wb.save(path)
    return path


# ===========================================================================
# EMAIL
# ===========================================================================
GMAIL_USER = os.environ.get("GMAIL_USER")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD")
ALERT_EMAIL_TO = os.environ.get("ALERT_EMAIL_TO")

# --- GA4: CTA banner do blog ---
GA4_PROPERTY_ID = os.environ.get("GA4_PROPERTY_ID", "")
CTA_UTM_MEDIUM = "cta_banner"

MAX_ITEMS_PER_SECTION = 8   # o e-mail mostra o topo; a planilha tem a lista inteira

GATILHO_INFO = [
    ("g1", "CTR Gap", "#FCE4D6"),
    ("g2", "Early Impression Drop", "#FFF2CC"),
    ("g3", "Emerging New Query", "#DDEEFF"),
    ("g4", "Optimized Post Decay", "#FCE4D6"),
    ("g5", "New Content Life Signal", "#DDEEFF"),
    ("g6", "City Pages (Rising/Falling)", "#FFF2CC"),
    ("g7", "Brand Traffic", "#FFF2CC"),
]


def _html_escape(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


# Guidance per trigger type, split into HYPOTHESIS (what might explain it -
# never certainty) and VERIFY (a 2-minute check before touching anything).
# None of the 8 triggers can see title tags, on-page content, or the live
# SERP - they only see aggregate GSC/GA4 numbers. A prescriptive "rewrite
# the title" is a guess dressed up as an instruction: the homepage might
# already be well-optimized for this exact intent, and the real cause could
# be a SERP feature, a better-suited page elsewhere on the site, or nothing
# fixable at all. Verify first; act only on what verification finds.
GUIDANCE_BY_TRIGGER = {
    "1. CTR Gap": {
        "hypothesis": "Three possible causes, not mutually exclusive: the title/meta may not match this query's intent as written; a SERP feature (map pack, People Also Ask, ads) may be taking the click above your result; or a different page on the site may be a more precise match for this exact query than the one ranking.",
        "verify": "Search this exact query in an incognito window. Check: (a) does the live title/snippet actually match it, (b) is there a map pack/PAA/ad block above position 5, (c) is the ranking page (see Page) really the best match, or would another page fit better.",
        "action": "Only rewrite the title/meta if (a) shows a real mismatch. If a SERP feature is the cause, there is no copy fix - track it instead. If a different page would fit better, that's a content/IA decision, not a copy edit.",
    },
    "2. Early Impression Drop": {
        "hypothesis": "Position held steady while impressions dropped, which usually points at Google showing this result less (a new AI Overview or SERP feature), not a ranking problem.",
        "verify": "Check the live SERP for this query for a new AI Overview, featured snippet, or other feature that wasn't there in the prior period.",
        "action": "No content or ranking action is likely needed. If a feature is confirmed, there is currently no fix - monitor it. If verification shows something else changed, treat as a new finding, not this one.",
    },
    "4. Optimized Post Decay": {
        "hypothesis": "A competitor may have published fresher or more complete content, or this post may no longer fully answer the query as well as it once did.",
        "verify": "Compare this page against the current top 3 results for its main query - what do they cover, or cover more recently, that this page doesn't?",
        "action": "Refresh the content only if verification shows a real gap versus what's currently ranking above it.",
    },
    "5. Content Life Signal (Negative)": {
        "hypothesis": "Most likely an indexing issue - the page may not have been crawled yet, or is blocked/excluded.",
        "verify": "Run URL Inspection in GSC for this exact page.",
        "action": "Follow whatever URL Inspection reports (request indexing, fix a block, etc.).",
    },
    "6. City Page Falling": {
        "hypothesis": "Could be a technical issue on the page itself, a change in the local map pack, or new competitor activity specific to this city.",
        "verify": "Load the page to confirm it works, check this city's map pack position, and scan the live SERP for a new competitor.",
        "action": "Act on whichever of the three verification finds - do not assume a content rewrite is the fix without checking first.",
    },
    "7. Brand - POSITION DROPPING": {
        "hypothesis": "Something is now outranking the homepage for the brand's own name - could be a competitor, a review/complaint site, or a directory listing.",
        "verify": "Search the brand query directly and see what outranks the homepage.",
        "action": "The fix depends entirely on what's found - there's no generic action for this until verification identifies the competing result.",
    },
    "7. Brand - Traffic Dropping": {
        "hypothesis": "Google may be showing an AI Overview, Knowledge Panel, or other feature in place of a click-through to the site for brand searches.",
        "verify": "Check the Knowledge Panel and the live SERP for the brand term for a new feature sitting above the organic result.",
        "action": "If a feature is confirmed, there's no direct fix - this is worth tracking. If nothing has changed on the SERP, treat as an open question, not a confirmed cause.",
    },
    "9. Cannibalization": {
        "hypothesis": "Two or more of your own pages are competing for the same query, which typically holds the best-ranking page back and confuses which page Google should show.",
        "verify": "Decide which page should own this query - usually the one already ranking best or most specific to the intent.",
        "action": "Consolidate the weaker page into the stronger one (301 redirect or merge content), or clearly differentiate them if both should legitimately exist.",
    },
}

URGENT_TRIGGER_TYPES = set(GUIDANCE_BY_TRIGGER.keys())


def _task_impact(alert):
    """
    A rough, comparable 'how much is at stake' number, used only to order
    tasks within the same priority tier - not to decide the tier itself.
    Impressions and clicks aren't on the same scale, so clicks get a
    multiplier; this doesn't need to be precise, just directionally right.
    """
    if alert.get("impressoes_semana") is not None:
        return float(alert["impressoes_semana"])
    if alert.get("media_clicks_4sem") is not None:
        return float(alert["media_clicks_4sem"]) * 10
    return 0.0


def build_urgent_tasks(g1, g2, g3, g4, g5, g6, g7, g8, g9=None, systemic=None):
    """
    Filters the 9 triggers down to only what represents a real problem -
    the subset Rafael turns into ClickUp tasks. Opportunity/good-news
    signals (new queries, pages rising, positive life signals, and G8 rows
    without a break) are left out of this list on purpose; they still live
    in the full spreadsheet.

    Each task carries hypothesis + verify + action rather than a single
    prescriptive instruction, because none of these triggers can see
    on-page content or the live SERP - only aggregate numbers.

    Ordering has two layers, not one: priority tier first (MAXIMUM > HIGH >
    MEDIUM), then within a tier, by how much is actually at stake
    (impressions/clicks). A documented known issue also gets pulled to the
    top of its tier - a confirmed, already-diagnosed problem is the most
    actionable item on the list, more useful to act on first than a bigger
    but unconfirmed one.

    If a systemic pattern was detected (many pages dropping in the same
    tight range, same week), the individual G2/G6 alerts for those pages
    are folded into ONE task instead of appearing as N near-duplicates -
    that was real noise found in production: 8 city pages each showing up
    as a separate "check this page" task when they were one event.
    """
    g9 = g9 or []
    tasks = []
    absorbed_pages = set(systemic["pages"]) if systemic else set()

    if systemic:
        tasks.append({
            "priority": "MAXIMUM",
            "_impact": systemic["total_impressions"] + 2_000_000,
            "trigger": systemic["gatilho"],
            "target": f"{systemic['n_pages']} pages ({systemic['drop_range']} impression drop)",
            "page": ", ".join(systemic["pages"][:5]) + (
                f" +{len(systemic['pages'])-5} more" if len(systemic["pages"]) > 5 else ""),
            "evidence": systemic["motivo"],
            "hypothesis": ("A single event most likely explains all of these at once - check "
                          "one query from this group live before treating this as several "
                          "separate problems."),
            "verify": ("Search one or two of the affected queries live. Look for a new AI "
                      "Overview, local pack change, or other feature that wasn't there before, "
                      "and confirm it appears across more than one of the affected queries."),
            "action": ("If confirmed as one event, there's likely no per-page fix - track it "
                      "as a single trend, not N separate tasks. If the affected queries turn "
                      "out unrelated on closer look, treat each page individually instead."),
        })

    for alert in g1 + g2 + g4 + g5 + g6 + g7 + g9:
        trigger = alert.get("gatilho")
        if trigger not in URGENT_TRIGGER_TYPES:
            continue
        if alert.get("actionable") is False:
            continue   # diagnosis already concluded there's no real fix - not a task
        if trigger in ("2. Early Impression Drop", "6. City Page Falling") and \
                norm_path(alert.get("pagina", "")) in absorbed_pages:
            continue   # already represented in the systemic task above

        priority = "MAXIMUM" if alert.get("urgencia") == "MAXIMUM" else (
            "HIGH" if alert.get("confianca") == "HIGH" or alert.get("urgencia") == "HIGH"
            else "MEDIUM")
        guidance = GUIDANCE_BY_TRIGGER[trigger]

        # G1 alerts enriched against the real SEO Log (keyword/title/known
        # issues) carry their own hypothesis - more reliable than the
        # generic multi-cause guess, because it's grounded in what the page
        # actually targets, not a guess about what it might target.
        hypothesis = alert.get("hypothesis_override") or guidance["hypothesis"]
        verify = alert.get("verify_override") or guidance["verify"]
        action = alert.get("action_override") or guidance["action"]

        # G4 (post decay): a documented known issue is more reliable than
        # the generic "compare against competitors" guess.
        has_known_issue = bool(alert.get("known_issue"))
        if trigger == "4. Optimized Post Decay" and has_known_issue:
            hypothesis = f"This page has a documented issue from its last optimization: \"{alert['known_issue']}\""
            verify = "Confirm whether this documented issue is still unresolved."
            action = "Fix the documented issue first - it is more likely the cause than external competition."
        elif "documented issue" in hypothesis:
            has_known_issue = True

        impact = _task_impact(alert)
        if has_known_issue:
            # a confirmed cause beats a bigger but unconfirmed one - float
            # to the top of its tier, and never leave it stuck at MEDIUM.
            if priority == "MEDIUM":
                priority = "HIGH"
            impact += 1_000_000

        tasks.append({
            "priority": priority,
            "_impact": impact,
            "trigger": trigger,
            "target": alert.get("query") or alert.get("pagina", ""),
            "page": alert.get("pagina", ""),
            "evidence": alert.get("motivo", ""),
            "hypothesis": hypothesis,
            "verify": verify,
            "action": action,
        })

    for r in g8:
        if not r.get("alerta"):
            continue
        tasks.append({
            "priority": "HIGH",
            "_impact": float(r.get("sessoes_semana_anterior", 0)) * 10,
            "trigger": "8. Blog CTA",
            "target": r["campanha"],
            "page": "",
            "evidence": r["alerta"],
            "hypothesis": "The CTA banner was removed from the post, the link broke, or the UTM parameter was dropped or changed.",
            "verify": "Open the post and confirm the banner and link are still there, and that the link still carries utm_medium=cta_banner.",
            "action": "Restore or fix whichever of the three verification finds broken.",
        })

    order = {"MAXIMUM": 0, "HIGH": 1, "MEDIUM": 2}
    tasks.sort(key=lambda t: (order.get(t["priority"], 3), -t["_impact"]))
    for i, t in enumerate(tasks, 1):
        t["num"] = i
    return tasks


PRIORITY_COLOR = {
    "MAXIMUM": ("#FFEBEE", "#D32F2F"),
    "HIGH": ("#FFF3E0", "#E65100"),
    "MEDIUM": ("#FFFDE7", "#F9A825"),
}


def _build_html_email(tasks, run_date, opportunity_count):
    """
    Rafael's ask: the email should arrive Monday with only what is urgent
    and needs to become a ClickUp task - not the full picture of all 8
    triggers. Everything else (new-query opportunities, pages rising,
    positive life signals) still lives in the attached spreadsheet, which
    he sends alongside the Monday SE Ranking audit.

    Each task carries a concrete next step (ACTION_BY_TRIGGER), so a task
    can go straight into ClickUp without being rewritten.
    """
    html = ['<div style="font-family:Arial,Helvetica,sans-serif;max-width:720px;">']
    html.append('<h2 style="color:#1F3864;margin-bottom:4px;">TurFresh - Weekly Tasks</h2>')
    html.append(f'<p style="color:#666;margin-top:0;">{run_date.isoformat()}</p>')

    if not tasks:
        html.append('<div style="border-radius:8px;padding:16px 20px;background:#E8F5E9;'
                    'border:1px solid #A5D6A7;font-family:Arial,Helvetica,sans-serif;">')
        html.append('<strong>No urgent tasks this week.</strong>')
        if opportunity_count:
            html.append(f'<p style="margin:8px 0 0 0;color:#555;">{opportunity_count} '
                       f'opportunity/visibility items (new queries, pages rising) are in '
                       f'the attached spreadsheet - no action needed, FYI only.</p>')
        html.append('</div></div>')
        return "\n".join(html)

    for t in tasks:
        bg, border = PRIORITY_COLOR.get(t["priority"], PRIORITY_COLOR["MEDIUM"])
        html.append(f'<div style="border-radius:8px;padding:14px 18px;margin-bottom:12px;'
                    f'background:{bg};border-left:4px solid {border};'
                    f'font-family:Arial,Helvetica,sans-serif;">')
        html.append(f'<div style="font-size:11px;font-weight:bold;color:{border};'
                    f'text-transform:uppercase;margin-bottom:4px;">'
                    f'#{t["num"]} · {t["priority"]} · {_html_escape(t["trigger"])}</div>')
        html.append(f'<div style="font-size:14px;font-weight:bold;margin-bottom:4px;">'
                    f'{_html_escape(t["target"])}</div>')
        if t["page"] and t["page"] != t["target"]:
            html.append(f'<div style="font-size:12px;color:#666;margin-bottom:6px;">'
                        f'{_html_escape(t["page"])}</div>')
        html.append(f'<div style="font-size:13px;color:#333;margin-bottom:5px;">'
                    f'<b>Evidence:</b> {_html_escape(t["evidence"])}</div>')
        html.append(f'<div style="font-size:13px;color:#333;margin-bottom:5px;">'
                    f'<b>Possible cause:</b> {_html_escape(t["hypothesis"])}</div>')
        html.append(f'<div style="font-size:13px;color:#333;margin-bottom:5px;">'
                    f'<b>Verify first:</b> {_html_escape(t["verify"])}</div>')
        html.append(f'<div style="font-size:13px;color:#333;">'
                    f'<b>Then:</b> {_html_escape(t["action"])}</div>')
        html.append('</div>')

    if opportunity_count:
        html.append(f'<p style="color:#666;font-size:12px;">Plus {opportunity_count} '
                   f'opportunity/visibility items (new queries, pages rising, positive '
                   f'signals) with no action needed - see the attached spreadsheet.</p>')
    html.append('<p style="color:#666;font-size:12px;">Full detail and evidence for every '
               'trigger, including the ones not listed above, in the attached spreadsheet.</p>')
    html.append('</div>')
    return "\n".join(html)


def send_email(g1, g2, g3, g4, g5, g6, g7, g8, g9, systemic, report_path, run_date):
    import smtplib
    from email.message import EmailMessage

    tasks = build_urgent_tasks(g1, g2, g3, g4, g5, g6, g7, g8, g9, systemic)

    # everything that did NOT become a task: opportunities and good news,
    # kept out of the email on purpose, still counted so Rafael knows the
    # spreadsheet has more context even when the email is short.
    total_all = (len(g1) + len(g2) + len(g3) + len(g4) + len(g5) + len(g6)
                + len(g7) + len(g8) + len(g9))
    opportunity_count = max(total_all - len(tasks), 0)

    maximum_tasks = [t for t in tasks if t["priority"] == "MAXIMUM"]

    # plain-text fallback, for email clients without HTML support
    text_lines = [f"TurFresh - Weekly Tasks - {run_date.isoformat()}", "=" * 55, ""]
    if not tasks:
        text_lines.append("No urgent tasks this week.")
        if opportunity_count:
            text_lines.append(f"{opportunity_count} opportunity/visibility items (no action "
                             f"needed) are in the attached spreadsheet.")
    else:
        for t in tasks:
            text_lines.append(f"#{t['num']} [{t['priority']}] {t['trigger']}")
            text_lines.append(f"    {t['target']}")
            if t["page"] and t["page"] != t["target"]:
                text_lines.append(f"    {t['page']}")
            text_lines.append(f"    Evidence: {t['evidence']}")
            text_lines.append(f"    Possible cause: {t['hypothesis']}")
            text_lines.append(f"    Verify first: {t['verify']}")
            text_lines.append(f"    Then: {t['action']}")
            text_lines.append("")
        if opportunity_count:
            text_lines.append(f"Plus {opportunity_count} opportunity/visibility items "
                             f"(no action needed) in the attached spreadsheet.")
    text_lines.append("\nFull detail for every trigger in the attached spreadsheet.")
    text = "\n".join(text_lines)

    html = _build_html_email(tasks, run_date, opportunity_count)

    if not (GMAIL_USER and GMAIL_APP_PASSWORD and ALERT_EMAIL_TO):
        print(text)
        return

    msg = EmailMessage()
    prefix = "[URGENT] " if maximum_tasks else ""
    task_word = "task" if len(tasks) == 1 else "tasks"
    msg["Subject"] = f"{prefix}[TurFresh] {len(tasks)} {task_word} - {run_date.isoformat()}"
    msg["From"] = GMAIL_USER
    msg["To"] = ALERT_EMAIL_TO
    msg.set_content(text)
    msg.add_alternative(html, subtype="html")
    with open(report_path, "rb") as f:
        msg.add_attachment(f.read(), maintype="application",
                           subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           filename=f"turfresh-alertas-{run_date.isoformat()}.xlsx")
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
        s.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        s.send_message(msg)
    print("Email enviado (HTML).")
    print(text)


# ===========================================================================
# GA4 - CTA BANNER DO BLOG (Gatilho 8)
# ===========================================================================
def get_ga4_client():
    """
    Falha graciosamente: se a API nao estiver ativada, o secret faltando, ou
    a conta de servico sem acesso, devolve None em vez de derrubar o script
    inteiro. Os outros 7 gatilhos (GSC) nao dependem do GA4 e devem continuar
    funcionando mesmo se essa parte falhar.

    Usa credenciais PROPRIAS (GA4_CLIENT_EMAIL/GA4_PRIVATE_KEY), diferentes
    das do GSC - essa conta de servico vive no projeto guiamexico-seo-alerts,
    nao no turfresh-seo-alerts, porque a conta pinsoda.com (que criou o
    projeto da TurFresh) e gerenciada por uma politica de Workspace que
    bloqueou a ativacao da API GA4 nesse projeto. Contornado usando um
    projeto que ja tem a API liberada.
    """
    if not GA4_PROPERTY_ID:
        print("  GA4_PROPERTY_ID nao configurado - pulando Gatilho 8 (CTA banner).")
        return None
    ga4_email = os.environ.get("GA4_CLIENT_EMAIL")
    ga4_key = os.environ.get("GA4_PRIVATE_KEY")
    if not ga4_email or not ga4_key:
        print("  GA4_CLIENT_EMAIL ou GA4_PRIVATE_KEY nao configurados - pulando Gatilho 8.")
        return None
    try:
        from google.analytics.data_v1beta import BetaAnalyticsDataClient
        creds = service_account.Credentials.from_service_account_info(
            {"type": "service_account", "client_email": ga4_email,
             "private_key": ga4_key,
             "token_uri": "https://oauth2.googleapis.com/token"},
            scopes=["https://www.googleapis.com/auth/analytics.readonly"])
        return BetaAnalyticsDataClient(credentials=creds)
    except Exception as e:
        print(f"  GA4 nao disponivel ({e}) - pulando Gatilho 8 (CTA banner).")
        return None


def fetch_cta_banner_by_campaign(client, start_date, end_date):
    """
    Sessoes chegando com utm_medium=cta_banner, agrupadas por utm_campaign
    (que no seu setup e o nome do post). Isso responde 'qual post esta
    levando gente pro contato', nao so 'quantos clicks no total'.
    """
    if client is None:
        return pd.DataFrame()
    try:
        from google.analytics.data_v1beta.types import (
            DateRange, Dimension, Metric, RunReportRequest, Filter, FilterExpression)
        request = RunReportRequest(
            property=f"properties/{GA4_PROPERTY_ID}",
            date_ranges=[DateRange(start_date=start_date.isoformat(),
                                   end_date=end_date.isoformat())],
            dimensions=[Dimension(name="sessionCampaignName")],
            metrics=[Metric(name="sessions")],
            dimension_filter=FilterExpression(
                filter=Filter(field_name="sessionMedium",
                             string_filter=Filter.StringFilter(value=CTA_UTM_MEDIUM))),
        )
        resp = client.run_report(request)
        rows = []
        for r in resp.rows:
            rows.append({"campanha": r.dimension_values[0].value,
                        "sessoes": int(r.metric_values[0].value)})
        return pd.DataFrame(rows)
    except Exception as e:
        print(f"  Erro ao buscar dados de CTA banner: {e}")
        return pd.DataFrame()


def gatilho_8_cta_banner(current_cta, previous_cta):
    """
    Nao e deteccao de anomalia como os outros 7 - e visibilidade direta de
    lead. Toda campanha (post) com sessao essa semana aparece, ordenada por
    volume. Alem disso, sinaliza quando uma campanha que tinha sessoes na
    semana anterior caiu para zero (CTA pode estar quebrado ou removido).
    """
    rows = []
    prev_map = dict(zip(previous_cta.get("campanha", []), previous_cta.get("sessoes", [])))

    for _, r in current_cta.iterrows():
        campanha, sessoes = r["campanha"], int(r["sessoes"])
        prev = prev_map.get(campanha, 0)
        if prev > 0:
            var = f"{(sessoes-prev)/prev*100:+.0f}%"
        else:
            var = "new"
        rows.append({"campanha": campanha, "sessoes_semana": sessoes,
                    "sessoes_semana_anterior": prev, "variacao": var,
                    "alerta": ""})

    # campanhas que tinham sessao e cairam pra zero essa semana
    current_campaigns = set(current_cta.get("campanha", []))
    for campanha, prev in prev_map.items():
        if prev >= 3 and campanha not in current_campaigns:
            rows.append({"campanha": campanha, "sessoes_semana": 0,
                        "sessoes_semana_anterior": prev, "variacao": "-100%",
                        "alerta": "Had sessions, dropped to zero - check if the CTA still exists on the post"})

    rows.sort(key=lambda x: -x["sessoes_semana"])
    return rows


# ===========================================================================
# MAIN

# ===========================================================================
def main():
    run_date = date.today()
    service = get_gsc_service()
    current, trailing, windows = fetch_trailing_weeks(service)

    print("Calculando medias moveis de 4 semanas...")
    stats = build_trailing_stats(trailing)
    print(f"  {len(stats)} combinacoes query+pagina no historico\n")

    print("Carregando dados de referencia...")
    seo_log_urls = load_seo_log_urls()
    city_pages = load_city_pages()
    print()

    print_funnel_diagnostic(current)

    print("Rodando os 9 gatilhos...")
    g1 = gatilho_1_vazamento_ctr(current, stats, seo_log_urls)
    g2 = gatilho_2_queda_precoce(current, stats)
    g3 = gatilho_3_query_nova(current, stats)
    g4 = gatilho_4_decaimento_posts(current, stats, seo_log_urls)
    g5 = gatilho_5_sinal_vida(current, city_pages, run_date)
    g6 = gatilho_6_city_pages(current, stats, city_pages)
    g7 = gatilho_7_marca(current, stats)
    g9 = gatilho_9_canibalizacao(current)

    all_alerts = g1 + g2 + g3 + g4 + g5 + g6 + g7 + g9
    print(f"  G1 vazamento CTR: {len(g1)}")
    print(f"  G2 queda precoce: {len(g2)}")
    print(f"  G3 query nova: {len(g3)}")
    print(f"  G4 decaimento posts: {len(g4)}")
    print(f"  G5 sinal de vida: {len(g5)}")
    print(f"  G6 city pages: {len(g6)}")
    print(f"  G7 marca: {len(g7)}")
    print(f"  G9 canibalizacao: {len(g9)}")
    print(f"  TOTAL: {len(all_alerts)} alertas\n")

    systemic = detect_systemic_pattern(g2, g6)
    if systemic:
        print(f"PADRAO SISTEMICO DETECTADO: {systemic['n_pages']} paginas, "
              f"faixa de queda {systemic['drop_range']}\n")

    radar = build_radar(current, stats, all_alerts)
    print(f"Radar: {len(radar)} paginas com volume real\n")

    print("Buscando dados de CTA banner (GA4)...")
    ga4_client = get_ga4_client()
    cur_start, cur_end = windows[-1]
    prev_start, prev_end = windows[-2]
    current_cta = fetch_cta_banner_by_campaign(ga4_client, cur_start, cur_end)
    previous_cta = fetch_cta_banner_by_campaign(ga4_client, prev_start, prev_end)
    g8 = gatilho_8_cta_banner(current_cta, previous_cta)
    print(f"  G8 CTA banner: {len(g8)} campanhas com dado\n")

    report_path = write_report(g1, g2, g3, g4, g5, g6, g7, g8, g9, systemic, radar, run_date)
    print(f"Relatorio: {report_path}\n")

    send_email(g1, g2, g3, g4, g5, g6, g7, g8, g9, systemic, report_path, run_date)


if __name__ == "__main__":
    main()
